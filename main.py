import os

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter


def process_file(file_path):
    data = {}
    with open(file_path, 'r') as file:
        for line in file:
            if ':' in line:
                key, value = line.strip().split(':', 1)
                data[key.strip()] = value.strip()
    return data


def get_database_and_isolation(folder, file_name):
    parts = file_name.split('_')
    database = folder.split('_')[0]
    isolation = parts[0].replace('-', ' ').title()
    return database, isolation


all_data = []
folders = ['MySQL_8.0.39', 'PostgreSQL_12.20']

for folder in folders:
    summary_folder = os.path.join(folder, 'result_summary')
    if os.path.exists(summary_folder):
        for filename in os.listdir(summary_folder):
            if filename.endswith('total-result.txt'):
                file_path = os.path.join(summary_folder, filename)
                data = process_file(file_path)
                database, isolation = get_database_and_isolation(folder, filename)
                for operation, handling in data.items():
                    all_data.append({
                        'Database': database,
                        'Isolation Level': isolation,
                        'Operation': operation,
                        'Handling': handling
                    })

df = pd.DataFrame(all_data)

pivot_df = df.pivot_table(
    values='Handling',
    index='Operation',
    columns=['Database', 'Isolation Level'],
    aggfunc='first'
)

output_filename = 'transaction_summary.xlsx'
pivot_df.to_excel(output_filename)

# 使用 openpyxl 调整单元格大小
wb = load_workbook(output_filename)
ws = wb.active

# 设置列宽
for column in ws.columns:
    max_length = 0
    column_letter = get_column_letter(column[0].column)
    for cell in column:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(cell.value)
        except:
            pass
    adjusted_width = (max_length + 2) * 1.2
    ws.column_dimensions[column_letter].width = adjusted_width

# 设置行高
for row in ws.rows:
    ws.row_dimensions[row[0].row].height = 20

# 设置所有单元格居中对齐
for row in ws.iter_rows():
    for cell in row:
        cell.alignment = Alignment(horizontal='center', vertical='center')

# 保存调整后的文件
wb.save(output_filename)

print(f"数据已保存到 '{output_filename}'")
